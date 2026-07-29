#!/usr/bin/env python3
"""更新本地举报台账中的一条已有记录。

用法：
    python scripts/update_ledger.py --seq 1 --updates update.json [--ledger ledger.xlsx]

只允许更新举报渠道、举报单号、举报状态和备注；不会新增记录。
"""
import argparse
import json
from pathlib import Path

from openpyxl import load_workbook


ALLOWED_FIELDS = {"举报渠道", "举报单号", "举报状态", "备注"}


def main():
    parser = argparse.ArgumentParser(description="更新本地举报台账中的已有记录")
    parser.add_argument("--ledger", default="ledger.xlsx", help="台账 Excel 文件路径")
    parser.add_argument("--seq", required=True, type=int, help="需要更新的记录序号")
    parser.add_argument("--updates", required=True, help="待更新字段的 JSON 文件路径")
    args = parser.parse_args()

    ledger_path = Path(args.ledger)
    updates = json.loads(Path(args.updates).read_text(encoding="utf-8"))
    if not isinstance(updates, dict):
        raise SystemExit("更新文件必须是 JSON 对象")
    if not updates:
        raise SystemExit("没有可更新的字段")
    disallowed_fields = set(updates) - ALLOWED_FIELDS
    if disallowed_fields:
        fields = "、".join(sorted(disallowed_fields))
        raise SystemExit(f"不允许更新以下字段：{fields}")

    wb = load_workbook(ledger_path)
    ws = wb.active
    headers = {
        cell.value: cell.column
        for cell in ws[1]
        if isinstance(cell.value, str)
    }

    matching_rows = [
        row
        for row in range(2, ws.max_row + 1)
        if ws.cell(row=row, column=headers["序号"]).value == args.seq
    ]
    if not matching_rows:
        raise SystemExit(f"未找到序号 {args.seq} 的台账记录")
    if len(matching_rows) > 1:
        raise SystemExit(f"序号 {args.seq} 存在重复记录，已停止更新")
    target_row = matching_rows[0]
    for field, value in updates.items():
        ws.cell(row=target_row, column=headers[field], value=value)

    wb.save(ledger_path)
    print(f"已更新台账：{ledger_path}，序号 {args.seq}")


if __name__ == "__main__":
    main()
