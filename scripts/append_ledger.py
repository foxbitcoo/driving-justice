#!/usr/bin/env python3
"""向本地举报台账（Excel）追加一条记录，并把证据关键帧图片嵌入到对应单元格。

用法：
    python scripts/append_ledger.py --entry entry.json [--ledger ledger.xlsx]

entry.json 字段说明见 templates/ledger-schema.md。
依赖：pip install openpyxl pillow
"""
import argparse
import json
import sys
from pathlib import Path

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.exit("缺少依赖，请先运行：pip install openpyxl pillow")

HEADERS = [
    "序号", "举报时间", "视频文件名", "发生地点", "违法行为", "法律依据",
    "车牌号", "车辆类型", "车辆颜色",
    "证据图1", "证据图2", "证据图3", "证据图4", "证据图5",
    "原始证据文件路径", "举报渠道", "举报单号", "举报状态", "备注",
]
IMAGE_COLS = ["证据图1", "证据图2", "证据图3", "证据图4", "证据图5"]
THUMB_WIDTH = 140
THUMB_HEIGHT = 100
IMAGE_ROW_HEIGHT = 78  # 单位：磅，约等于 THUMB_HEIGHT 像素换算


def load_or_create(ledger_path: Path):
    if ledger_path.exists():
        wb = load_workbook(ledger_path)
        ws = wb.active
        return wb, ws
    wb = Workbook()
    ws = wb.active
    ws.title = "台账"
    ws.append(HEADERS)
    for col in IMAGE_COLS:
        idx = HEADERS.index(col) + 1
        ws.column_dimensions[get_column_letter(idx)].width = 20
    return wb, ws


def next_seq(ws) -> int:
    last = ws.max_row
    if last <= 1:
        return 1
    val = ws.cell(row=last, column=1).value
    try:
        return int(val) + 1
    except (TypeError, ValueError):
        return last  # 表头之外还有数据但序号异常时，兜底用当前行数


def main():
    parser = argparse.ArgumentParser(description="追加一条举报台账记录（含嵌入图片）")
    parser.add_argument("--entry", required=True, help="记录数据的 JSON 文件路径")
    parser.add_argument("--ledger", default="ledger.xlsx", help="台账 Excel 文件路径，默认 ledger.xlsx")
    args = parser.parse_args()

    entry = json.loads(Path(args.entry).read_text(encoding="utf-8"))
    ledger_path = Path(args.ledger)

    wb, ws = load_or_create(ledger_path)
    row_idx = ws.max_row + 1
    seq = next_seq(ws)

    image_paths = entry.get("证据图片", [])
    row_values = [
        seq,
        entry.get("举报时间", ""),
        entry.get("视频文件名", ""),
        entry.get("发生地点", ""),
        entry.get("违法行为", ""),
        entry.get("法律依据", ""),
        entry.get("车牌号", ""),
        entry.get("车辆类型", ""),
        entry.get("车辆颜色", ""),
        "", "", "", "", "",  # 证据图1~5 留空，图片单独插入
        "；".join(image_paths),
        entry.get("举报渠道", ""),
        entry.get("举报单号", ""),
        entry.get("举报状态", ""),
        entry.get("备注", ""),
    ]
    ws.append(row_values)
    ws.row_dimensions[row_idx].height = IMAGE_ROW_HEIGHT

    missing = []
    for i, img_path in enumerate(image_paths[:5]):
        p = Path(img_path)
        if not p.exists():
            missing.append(img_path)
            continue
        col_idx = HEADERS.index(IMAGE_COLS[i]) + 1
        img = XLImage(str(p))
        img.width = THUMB_WIDTH
        img.height = THUMB_HEIGHT
        cell_ref = f"{get_column_letter(col_idx)}{row_idx}"
        ws.add_image(img, cell_ref)

    wb.save(ledger_path)

    print(f"已写入台账：{ledger_path}，第 {row_idx} 行，序号 {seq}")
    if missing:
        print("以下图片路径不存在，未能嵌入，请检查：")
        for m in missing:
            print(f"  - {m}")
    if len(image_paths) > 5:
        print(f"注意：本条记录有 {len(image_paths)} 张图片，仅前 5 张被嵌入，其余路径仍保存在'原始证据文件路径'列中。")


if __name__ == "__main__":
    main()
