#!/usr/bin/env python3
"""把候选交通事件写入单个 Excel 结果文件，并嵌入证据图片。

车牌只能来自上游已通过结构化清晰度、轨迹和帧引用校验的模型判断。
"""

import argparse
import json
import sys
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.exit("缺少依赖，请先运行：pip install openpyxl pillow")


HEADERS = [
    "序号",
    "素材文件",
    "开始时间",
    "结束时间",
    "候选违法行为",
    "车辆类型",
    "车辆颜色",
    "车牌号",
    "证据状态",
    "证据说明",
    "证据图1",
    "证据图2",
    "证据图3",
    "证据图4",
    "证据图5",
    "证据帧标识",
    "模型链路",
    "备注",
]
IMAGE_HEADERS = [f"证据图{i}" for i in range(1, 6)]
THUMB_WIDTH = 160
THUMB_HEIGHT = 100


def seconds_to_timestamp(value) -> str:
    try:
        total_ms = max(0, round(float(value) * 1000))
    except (TypeError, ValueError):
        return ""
    hours, remainder = divmod(total_ms, 3_600_000)
    minutes, remainder = divmod(remainder, 60_000)
    seconds, milliseconds = divmod(remainder, 1000)
    return f"{hours:02d}:{minutes:02d}:{seconds:02d}.{milliseconds:03d}"


def load_json(path: Path) -> dict:
    data = json.loads(path.read_text(encoding="utf-8"))
    if not isinstance(data, dict):
        raise ValueError(f"{path} 必须是 JSON 对象")
    return data


def accepted_model_plate(event: dict) -> tuple[str, str]:
    if event.get("plate_fully_clear") is not True:
        return "", ""
    if event.get("trajectory_continuous") is not True:
        return "", ""
    plate = str(event.get("plate_number", "")).strip()
    side = str(event.get("plate_side", "")).strip()
    return (plate, side) if plate else ("", "")


def write_workbook(
    analysis: dict,
    output_path: Path,
) -> None:
    events = analysis.get("events", [])
    if not isinstance(events, list):
        raise ValueError("analysis.events 必须是数组")

    wb = Workbook()
    ws = wb.active
    ws.title = "识别结果"
    ws.append(HEADERS)

    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in ws[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}1"

    widths = {
        "A": 8,
        "B": 28,
        "C": 15,
        "D": 15,
        "E": 24,
        "F": 14,
        "G": 12,
        "H": 16,
        "I": 20,
        "J": 36,
        "P": 48,
        "Q": 34,
        "R": 30,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    for image_header in IMAGE_HEADERS:
        col = get_column_letter(HEADERS.index(image_header) + 1)
        ws.column_dimensions[col].width = 22

    if not events:
        ws.append([
            1,
            analysis.get("source_summary", ""),
            "",
            "",
            "未发现候选事件",
            "",
            "",
            "",
            "无候选结果",
            "本次自动扫描未发现候选；不代表素材中一定没有违法。",
        ])
        ws.row_dimensions[2].height = 40
    else:
        for index, event in enumerate(events, start=1):
            if not isinstance(event, dict):
                continue
            plate, plate_side = accepted_model_plate(event)
            evidence_images = [
                str(path)
                for path in event.get("evidence_images", [])
                if isinstance(path, str) and path.strip()
            ]
            evidence_status = str(event.get("evidence_status", "待人工复核"))
            if plate:
                evidence_status = "模型判定车牌与轨迹通过（待人工复核）"
            row = [
                index,
                event.get("source_file", ""),
                seconds_to_timestamp(event.get("start_seconds")),
                seconds_to_timestamp(event.get("end_seconds")),
                event.get("behavior", ""),
                event.get("vehicle_type", ""),
                event.get("vehicle_color", ""),
                plate,
                evidence_status,
                event.get("evidence_note", ""),
                "",
                "",
                "",
                "",
                "",
                "；".join(
                    str(value)
                    for value in event.get("evidence_refs", [])
                    if isinstance(value, str)
                ),
                analysis.get("model_route", ""),
                (
                    f"模型识别{plate_side}牌；正式举报前仍需用户确认"
                    if plate_side
                    else event.get("notes", "")
                ),
            ]
            ws.append(row)
            row_index = ws.max_row
            ws.row_dimensions[row_index].height = 78
            for offset, image_path in enumerate(evidence_images[:5]):
                path = Path(image_path)
                if not path.is_file():
                    continue
                image_col = HEADERS.index(IMAGE_HEADERS[offset]) + 1
                image = XLImage(str(path))
                image.width = THUMB_WIDTH
                image.height = THUMB_HEIGHT
                anchor = f"{get_column_letter(image_col)}{row_index}"
                ws.add_image(image, anchor)

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def main() -> None:
    parser = argparse.ArgumentParser(description="生成嵌入证据图片的候选识别 Excel")
    parser.add_argument("--analysis", required=True, help="规范化候选事件 JSON")
    parser.add_argument("--output", default="analysis.xlsx", help="输出 Excel 路径")
    args = parser.parse_args()

    analysis = load_json(Path(args.analysis))
    output_path = Path(args.output)
    write_workbook(analysis, output_path)
    print(f"已生成识别结果：{output_path}")


if __name__ == "__main__":
    main()
