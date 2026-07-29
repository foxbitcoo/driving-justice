import base64
import json
import subprocess
import sys
import tempfile
import unittest
import zipfile
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as XLImage


PROJECT_ROOT = Path(__file__).resolve().parents[1]
UPDATE_SCRIPT = PROJECT_ROOT / "scripts" / "update_ledger.py"


class UpdateLedgerCliTests(unittest.TestCase):
    def test_updates_existing_row_without_adding_a_duplicate(self):
        with tempfile.TemporaryDirectory() as tmp_dir:
            tmp_path = Path(tmp_dir)
            ledger_path = tmp_path / "ledger.xlsx"
            updates_path = tmp_path / "updates.json"

            wb = Workbook()
            ws = wb.active
            ws.title = "台账"
            ws.append(["序号", "举报渠道", "举报单号", "举报状态", "备注"])
            ws.append([1, "珠海12345在线举报", "", "已预填待用户提交", ""])
            wb.save(ledger_path)

            updates_path.write_text(
                json.dumps(
                    {
                        "举报单号": "ZH202607290001",
                        "举报状态": "已提交待处理",
                    },
                    ensure_ascii=False,
                ),
                encoding="utf-8",
            )

            result = subprocess.run(
                [
                    sys.executable,
                    str(UPDATE_SCRIPT),
                    "--ledger",
                    str(ledger_path),
                    "--seq",
                    "1",
                    "--updates",
                    str(updates_path),
                ],
                capture_output=True,
                text=True,
            )

            self.assertEqual(result.returncode, 0, result.stderr)
            updated_wb = load_workbook(ledger_path)
            updated_ws = updated_wb.active
            self.assertEqual(updated_ws.max_row, 2)
            self.assertEqual(updated_ws["C2"].value, "ZH202607290001")
            self.assertEqual(updated_ws["D2"].value, "已提交待处理")

    def test_rejects_fields_outside_the_post_submission_allowlist(self):
        with tempfile.TemporaryDirectory() as tmp_dir:
            tmp_path = Path(tmp_dir)
            ledger_path = tmp_path / "ledger.xlsx"
            updates_path = tmp_path / "updates.json"

            wb = Workbook()
            ws = wb.active
            ws.title = "台账"
            ws.append(["序号", "车牌号", "举报状态"])
            ws.append([1, "粤C·12345", "已预填待用户提交"])
            wb.save(ledger_path)

            updates_path.write_text(
                json.dumps({"车牌号": "粤C·99999"}, ensure_ascii=False),
                encoding="utf-8",
            )

            result = subprocess.run(
                [
                    sys.executable,
                    str(UPDATE_SCRIPT),
                    "--ledger",
                    str(ledger_path),
                    "--seq",
                    "1",
                    "--updates",
                    str(updates_path),
                ],
                capture_output=True,
                text=True,
            )

            self.assertNotEqual(result.returncode, 0)
            unchanged_wb = load_workbook(ledger_path)
            self.assertEqual(unchanged_wb.active["B2"].value, "粤C·12345")

    def test_missing_sequence_reports_an_error_without_changing_the_ledger(self):
        with tempfile.TemporaryDirectory() as tmp_dir:
            tmp_path = Path(tmp_dir)
            ledger_path = tmp_path / "ledger.xlsx"
            updates_path = tmp_path / "updates.json"

            wb = Workbook()
            ws = wb.active
            ws.title = "台账"
            ws.append(["序号", "举报状态"])
            ws.append([1, "已预填待用户提交"])
            wb.save(ledger_path)

            updates_path.write_text(
                json.dumps({"举报状态": "已提交待处理"}, ensure_ascii=False),
                encoding="utf-8",
            )

            result = subprocess.run(
                [
                    sys.executable,
                    str(UPDATE_SCRIPT),
                    "--ledger",
                    str(ledger_path),
                    "--seq",
                    "99",
                    "--updates",
                    str(updates_path),
                ],
                capture_output=True,
                text=True,
            )

            self.assertNotEqual(result.returncode, 0)
            self.assertIn("未找到序号 99", result.stderr)
            unchanged_wb = load_workbook(ledger_path)
            self.assertEqual(unchanged_wb.active["B2"].value, "已预填待用户提交")

    def test_duplicate_sequence_reports_an_error_without_changing_either_row(self):
        with tempfile.TemporaryDirectory() as tmp_dir:
            tmp_path = Path(tmp_dir)
            ledger_path = tmp_path / "ledger.xlsx"
            updates_path = tmp_path / "updates.json"

            wb = Workbook()
            ws = wb.active
            ws.title = "台账"
            ws.append(["序号", "举报状态"])
            ws.append([1, "已预填待用户提交"])
            ws.append([1, "已预填待用户提交"])
            wb.save(ledger_path)

            updates_path.write_text(
                json.dumps({"举报状态": "已提交待处理"}, ensure_ascii=False),
                encoding="utf-8",
            )

            result = subprocess.run(
                [
                    sys.executable,
                    str(UPDATE_SCRIPT),
                    "--ledger",
                    str(ledger_path),
                    "--seq",
                    "1",
                    "--updates",
                    str(updates_path),
                ],
                capture_output=True,
                text=True,
            )

            self.assertNotEqual(result.returncode, 0)
            self.assertIn("序号 1 存在重复记录", result.stderr)
            unchanged_wb = load_workbook(ledger_path)
            unchanged_ws = unchanged_wb.active
            self.assertEqual(unchanged_ws["B2"].value, "已预填待用户提交")
            self.assertEqual(unchanged_ws["B3"].value, "已预填待用户提交")

    def test_empty_updates_are_rejected_instead_of_reporting_success(self):
        with tempfile.TemporaryDirectory() as tmp_dir:
            tmp_path = Path(tmp_dir)
            ledger_path = tmp_path / "ledger.xlsx"
            updates_path = tmp_path / "updates.json"

            wb = Workbook()
            ws = wb.active
            ws.title = "台账"
            ws.append(["序号", "举报状态"])
            ws.append([1, "已预填待用户提交"])
            wb.save(ledger_path)
            updates_path.write_text("{}", encoding="utf-8")

            result = subprocess.run(
                [
                    sys.executable,
                    str(UPDATE_SCRIPT),
                    "--ledger",
                    str(ledger_path),
                    "--seq",
                    "1",
                    "--updates",
                    str(updates_path),
                ],
                capture_output=True,
                text=True,
            )

            self.assertNotEqual(result.returncode, 0)
            self.assertIn("没有可更新的字段", result.stderr)

    def test_updating_submission_fields_keeps_embedded_evidence_images(self):
        with tempfile.TemporaryDirectory() as tmp_dir:
            tmp_path = Path(tmp_dir)
            ledger_path = tmp_path / "ledger.xlsx"
            updates_path = tmp_path / "updates.json"
            evidence_path = tmp_path / "evidence.png"
            evidence_path.write_bytes(
                base64.b64decode(
                    "iVBORw0KGgoAAAANSUhEUgAAAAEAAAABCAQAAAC1HAwCAAAAC0l"
                    "EQVR42mNk+A8AAQUBAScY42YAAAAASUVORK5CYII="
                )
            )

            wb = Workbook()
            ws = wb.active
            ws.title = "台账"
            ws.append(["序号", "举报状态", "证据图1"])
            ws.append([1, "已预填待用户提交", ""])
            ws.add_image(XLImage(str(evidence_path)), "C2")
            wb.save(ledger_path)

            with zipfile.ZipFile(ledger_path) as archive:
                media_before = {
                    name for name in archive.namelist() if name.startswith("xl/media/")
                }

            updates_path.write_text(
                json.dumps({"举报状态": "已提交待处理"}, ensure_ascii=False),
                encoding="utf-8",
            )
            result = subprocess.run(
                [
                    sys.executable,
                    str(UPDATE_SCRIPT),
                    "--ledger",
                    str(ledger_path),
                    "--seq",
                    "1",
                    "--updates",
                    str(updates_path),
                ],
                capture_output=True,
                text=True,
            )

            self.assertEqual(result.returncode, 0, result.stderr)
            with zipfile.ZipFile(ledger_path) as archive:
                media_after = {
                    name for name in archive.namelist() if name.startswith("xl/media/")
                }
            self.assertEqual(media_after, media_before)
            self.assertEqual(len(media_after), 1)


if __name__ == "__main__":
    unittest.main()
