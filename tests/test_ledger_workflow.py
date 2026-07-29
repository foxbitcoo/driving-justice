import base64
import json
import subprocess
import sys
import tempfile
import unittest
import zipfile
from pathlib import Path

from openpyxl import load_workbook


PROJECT_ROOT = Path(__file__).resolve().parents[1]
APPEND_SCRIPT = PROJECT_ROOT / "scripts" / "append_ledger.py"
UPDATE_SCRIPT = PROJECT_ROOT / "scripts" / "update_ledger.py"


class LedgerWorkflowTests(unittest.TestCase):
    def test_prefill_record_can_be_updated_after_manual_submission(self):
        with tempfile.TemporaryDirectory() as tmp_dir:
            tmp_path = Path(tmp_dir)
            ledger_path = tmp_path / "ledger.xlsx"
            entry_path = tmp_path / "entry.json"
            updates_path = tmp_path / "updates.json"
            evidence_path = tmp_path / "test-evidence.png"

            evidence_path.write_bytes(
                base64.b64decode(
                    "iVBORw0KGgoAAAANSUhEUgAAAAEAAAABCAQAAAC1HAwCAAAAC0l"
                    "EQVR42mNk+A8AAQUBAScY42YAAAAASUVORK5CYII="
                )
            )
            entry_path.write_text(
                json.dumps(
                    {
                        "举报时间": "2026-07-29 12:00",
                        "视频文件名": "test-only.mp4",
                        "发生地点": "测试地点（非真实举报）",
                        "违法行为": "测试行为（非真实举报）",
                        "法律依据": "测试依据（非真实举报）",
                        "车牌号": "TEST-PLATE",
                        "车辆类型": "测试车辆",
                        "车辆颜色": "测试颜色",
                        "证据图片": [str(evidence_path)],
                        "举报渠道": "测试渠道",
                        "举报单号": "",
                        "举报状态": "已预填待用户提交",
                        "备注": "自动化测试数据",
                    },
                    ensure_ascii=False,
                ),
                encoding="utf-8",
            )

            append_result = subprocess.run(
                [
                    sys.executable,
                    str(APPEND_SCRIPT),
                    "--entry",
                    str(entry_path),
                    "--ledger",
                    str(ledger_path),
                ],
                capture_output=True,
                text=True,
            )
            self.assertEqual(append_result.returncode, 0, append_result.stderr)

            updates_path.write_text(
                json.dumps(
                    {
                        "举报单号": "TEST-RECEIPT-001",
                        "举报状态": "已提交待处理",
                    },
                    ensure_ascii=False,
                ),
                encoding="utf-8",
            )
            update_result = subprocess.run(
                [
                    sys.executable,
                    str(UPDATE_SCRIPT),
                    "--ledger",
                    str(ledger_path),
                    "--seq",
                    "1",
                    "--updates",
                    str(updates_path),
                ],
                capture_output=True,
                text=True,
            )
            self.assertEqual(update_result.returncode, 0, update_result.stderr)

            wb = load_workbook(ledger_path)
            ws = wb.active
            headers = {
                cell.value: cell.column
                for cell in ws[1]
                if isinstance(cell.value, str)
            }
            self.assertEqual(ws.max_row, 2)
            self.assertEqual(ws.cell(2, headers["序号"]).value, 1)
            self.assertEqual(
                ws.cell(2, headers["举报单号"]).value,
                "TEST-RECEIPT-001",
            )
            self.assertEqual(
                ws.cell(2, headers["举报状态"]).value,
                "已提交待处理",
            )

            with zipfile.ZipFile(ledger_path) as archive:
                media_files = [
                    name
                    for name in archive.namelist()
                    if name.startswith("xl/media/")
                ]
            self.assertEqual(len(media_files), 1)


if __name__ == "__main__":
    unittest.main()
