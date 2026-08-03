import base64
import json
import os
import shutil
import subprocess
import sys
import tempfile
import threading
import unittest
import zipfile
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path

from openpyxl import load_workbook


PROJECT_ROOT = Path(__file__).resolve().parents[1]
PIPELINE_SCRIPT = PROJECT_ROOT / "scripts" / "run_low_cost_pipeline.py"
PNG_1X1 = base64.b64decode(
    "iVBORw0KGgoAAAANSUhEUgAAAAEAAAABCAQAAAC1HAwCAAAAC0l"
    "EQVR42mNk+A8AAQUBAScY42YAAAAASUVORK5CYII="
)


def header_map(ws):
    return {
        cell.value: cell.column
        for cell in ws[1]
        if isinstance(cell.value, str)
    }


class JsonApiServer:
    def __init__(self, responder):
        self.requests = []
        owner = self

        class Handler(BaseHTTPRequestHandler):
            def do_POST(self):
                length = int(self.headers.get("Content-Length", "0"))
                payload = json.loads(self.rfile.read(length).decode("utf-8"))
                owner.requests.append({"path": self.path, "payload": payload})
                body = json.dumps(responder(payload), ensure_ascii=False).encode("utf-8")
                self.send_response(200)
                self.send_header("Content-Type", "application/json")
                self.send_header("Content-Length", str(len(body)))
                self.end_headers()
                self.wfile.write(body)

            def log_message(self, format, *args):
                return

        self.server = ThreadingHTTPServer(("127.0.0.1", 0), Handler)
        self.thread = threading.Thread(target=self.server.serve_forever, daemon=True)

    @property
    def base_url(self):
        host, port = self.server.server_address
        return f"http://{host}:{port}"

    def __enter__(self):
        self.thread.start()
        return self

    def __exit__(self, exc_type, exc_value, traceback):
        self.server.shutdown()
        self.server.server_close()
        self.thread.join(timeout=5)


class AnalysisPipelineTests(unittest.TestCase):
    def make_image(self, directory: Path) -> Path:
        image = directory / "evidence.png"
        image.write_bytes(PNG_1X1)
        return image

    def test_offline_json_creates_one_excel_and_drops_model_plate(self):
        with tempfile.TemporaryDirectory() as tmp_dir:
            tmp_path = Path(tmp_dir)
            image = self.make_image(tmp_path)
            normalized = tmp_path / "normalized.json"
            output = tmp_path / "analysis.xlsx"
            normalized.write_text(
                json.dumps(
                    {
                        "events": [
                            {
                                "event_id": "model-event",
                                "source_file": "test.png",
                                "start_seconds": 1.25,
                                "end_seconds": 2.5,
                                "behavior": "逆行",
                                "vehicle_type": "轿车",
                                "vehicle_color": "黑色",
                                "plate_number": "粤B·12345",
                                "evidence_status": "待人工复核",
                                "evidence_note": "模型声称车牌是粤B12345",
                                "evidence_images": [str(image)],
                            }
                        ]
                    },
                    ensure_ascii=False,
                ),
                encoding="utf-8",
            )

            result = subprocess.run(
                [
                    sys.executable,
                    str(PIPELINE_SCRIPT),
                    "--input",
                    str(image),
                    "--normalized-json",
                    str(normalized),
                    "--output",
                    str(output),
                ],
                capture_output=True,
                text=True,
            )
            self.assertEqual(result.returncode, 0, result.stderr)
            self.assertTrue(output.is_file())

            wb = load_workbook(output)
            ws = wb["识别结果"]
            headers = header_map(ws)
            self.assertEqual(ws.max_row, 2)
            self.assertEqual(ws.cell(2, headers["候选违法行为"]).value, "逆行")
            self.assertIsNone(ws.cell(2, headers["车牌号"]).value)
            self.assertIn(
                "[车牌字符已移除]",
                ws.cell(2, headers["证据说明"]).value,
            )
            with zipfile.ZipFile(output) as archive:
                media = [
                    name for name in archive.namelist() if name.startswith("xl/media/")
                ]
            self.assertEqual(len(media), 1)

    def test_plate_is_written_only_when_model_gate_and_frame_refs_pass(self):
        with tempfile.TemporaryDirectory() as tmp_dir:
            tmp_path = Path(tmp_dir)
            image = self.make_image(tmp_path)
            normalized = tmp_path / "normalized.json"
            output = tmp_path / "analysis.xlsx"
            normalized.write_text(
                json.dumps(
                    {
                        "events": [
                            {
                                "source_file": "test.png",
                                "behavior": "逆行",
                                "evidence_images": [str(image)],
                                "evidence_frame_ids": ["image-001"],
                                "model_plate_decision": {
                                    "plate_number": "粤B·12345",
                                    "plate_side": "rear",
                                    "plate_fully_clear": True,
                                    "trajectory_continuous": True,
                                    "plate_frame_ids": ["image-001"],
                                    "incident_frame_ids": ["image-001"],
                                    "independent_consensus": {
                                        "passed": True,
                                        "plate_number": "粤B12345",
                                        "plate_side": "rear",
                                        "frame_ids": ["image-001"],
                                    },
                                },
                            }
                        ]
                    },
                    ensure_ascii=False,
                ),
                encoding="utf-8",
            )

            result = subprocess.run(
                [
                    sys.executable,
                    str(PIPELINE_SCRIPT),
                    "--input",
                    str(image),
                    "--normalized-json",
                    str(normalized),
                    "--output",
                    str(output),
                ],
                capture_output=True,
                text=True,
            )
            self.assertEqual(result.returncode, 0, result.stderr)
            ws = load_workbook(output)["识别结果"]
            headers = header_map(ws)
            self.assertEqual(ws.cell(2, headers["车牌号"]).value, "粤B12345")
            self.assertEqual(
                ws.cell(2, headers["证据状态"]).value,
                "模型判定车牌与轨迹通过（待人工复核）",
            )

    def test_model_plate_is_dropped_when_trajectory_or_refs_fail(self):
        with tempfile.TemporaryDirectory() as tmp_dir:
            tmp_path = Path(tmp_dir)
            image = self.make_image(tmp_path)
            normalized = tmp_path / "normalized.json"
            output = tmp_path / "analysis.xlsx"
            normalized.write_text(
                json.dumps(
                    {
                        "events": [
                            {
                                "source_file": "test.png",
                                "behavior": "压实线/压导流线",
                                "model_plate_decision": {
                                    "plate_number": "粤B12345",
                                    "plate_side": "rear",
                                    "plate_fully_clear": True,
                                    "trajectory_continuous": False,
                                    "plate_frame_ids": ["missing-frame"],
                                    "incident_frame_ids": ["image-001"],
                                },
                            }
                        ]
                    },
                    ensure_ascii=False,
                ),
                encoding="utf-8",
            )
            result = subprocess.run(
                [
                    sys.executable,
                    str(PIPELINE_SCRIPT),
                    "--input",
                    str(image),
                    "--normalized-json",
                    str(normalized),
                    "--output",
                    str(output),
                ],
                capture_output=True,
                text=True,
            )
            self.assertEqual(result.returncode, 0, result.stderr)
            ws = load_workbook(output)["识别结果"]
            headers = header_map(ws)
            self.assertIsNone(ws.cell(2, headers["车牌号"]).value)

    def test_model_plate_is_dropped_when_independent_read_disagrees(self):
        with tempfile.TemporaryDirectory() as tmp_dir:
            tmp_path = Path(tmp_dir)
            image = self.make_image(tmp_path)
            normalized = tmp_path / "normalized.json"
            output = tmp_path / "analysis.xlsx"
            normalized.write_text(
                json.dumps(
                    {
                        "events": [
                            {
                                "source_file": "test.png",
                                "behavior": "违规变道",
                                "model_plate_decision": {
                                    "plate_number": "粤B12345",
                                    "plate_side": "rear",
                                    "plate_fully_clear": True,
                                    "trajectory_continuous": True,
                                    "plate_frame_ids": ["image-001"],
                                    "incident_frame_ids": ["image-001"],
                                    "independent_consensus": {
                                        "passed": True,
                                        "plate_number": "粤B12346",
                                        "plate_side": "rear",
                                        "frame_ids": ["image-001"],
                                    },
                                },
                            }
                        ]
                    },
                    ensure_ascii=False,
                ),
                encoding="utf-8",
            )
            result = subprocess.run(
                [
                    sys.executable,
                    str(PIPELINE_SCRIPT),
                    "--input",
                    str(image),
                    "--normalized-json",
                    str(normalized),
                    "--output",
                    str(output),
                ],
                capture_output=True,
                text=True,
            )
            self.assertEqual(result.returncode, 0, result.stderr)
            ws = load_workbook(output)["识别结果"]
            headers = header_map(ws)
            self.assertIsNone(ws.cell(2, headers["车牌号"]).value)

    def test_provider_chain_uses_lite_then_flash_and_pro_on_conflict(self):
        def vision_response(payload):
            content = payload["messages"][0]["content"]
            frame_text = next(
                item["text"]
                for item in content
                if item.get("type") == "text" and item.get("text", "").startswith("frame_id=")
            )
            frame_id = frame_text.split(";", 1)[0].split("=", 1)[1]
            result = {
                "candidates": [
                    {
                        "source_file": "evidence.png",
                        "start_seconds": 0,
                        "end_seconds": 0,
                        "behavior": "违停",
                        "vehicle_type": "轿车",
                        "vehicle_color": "白色",
                        "plate_number": "粤B12345",
                        "evidence_status": "待人工复核",
                        "evidence_note": "单张图只能初筛",
                        "evidence_frame_ids": [frame_id],
                    }
                ]
            }
            return {"choices": [{"message": {"content": json.dumps(result, ensure_ascii=False)}}]}

        def deepseek_response(payload):
            model = payload["model"]
            review_required = model == "deepseek-v4-flash"
            result = {
                "review_required": review_required,
                "events": [
                    {
                        "source_file": "evidence.png",
                        "start_seconds": 0,
                        "end_seconds": 0,
                        "behavior": "违停",
                        "vehicle_type": "轿车",
                        "vehicle_color": "白色",
                        "evidence_status": "证据不足",
                        "evidence_note": "缺少连续时序",
                        "evidence_frame_ids": ["image-001"],
                    }
                ],
            }
            return {"choices": [{"message": {"content": json.dumps(result, ensure_ascii=False)}}]}

        with tempfile.TemporaryDirectory() as tmp_dir:
            tmp_path = Path(tmp_dir)
            image = self.make_image(tmp_path)
            output = tmp_path / "analysis.xlsx"
            with JsonApiServer(vision_response) as vision_server, JsonApiServer(
                deepseek_response
            ) as deepseek_server:
                env = os.environ.copy()
                env["ARK_API_KEY"] = "test-ark-key"
                env["DEEPSEEK_API_KEY"] = "test-deepseek-key"
                result = subprocess.run(
                    [
                        sys.executable,
                        str(PIPELINE_SCRIPT),
                        "--input",
                        str(image),
                        "--output",
                        str(output),
                        "--ark-base-url",
                        vision_server.base_url,
                        "--deepseek-base-url",
                        deepseek_server.base_url,
                    ],
                    capture_output=True,
                    text=True,
                    env=env,
                )
            self.assertEqual(result.returncode, 0, result.stderr)
            self.assertEqual(
                vision_server.requests[0]["payload"]["model"],
                "doubao-seed-2-0-lite-260215",
            )
            self.assertEqual(
                [request["payload"]["model"] for request in deepseek_server.requests],
                ["deepseek-v4-flash", "deepseek-v4-pro"],
            )
            ws = load_workbook(output)["识别结果"]
            headers = header_map(ws)
            self.assertIsNone(ws.cell(2, headers["车牌号"]).value)
            self.assertIn(
                "deepseek-v4-pro",
                ws.cell(2, headers["模型链路"]).value,
            )

    @unittest.skipUnless(shutil.which("ffmpeg"), "ffmpeg is not installed")
    def test_video_input_is_sampled_and_embedded_in_offline_mode(self):
        with tempfile.TemporaryDirectory() as tmp_dir:
            tmp_path = Path(tmp_dir)
            video = tmp_path / "sample.mp4"
            normalized = tmp_path / "normalized.json"
            output = tmp_path / "analysis.xlsx"
            create_video = subprocess.run(
                [
                    shutil.which("ffmpeg"),
                    "-hide_banner",
                    "-loglevel",
                    "error",
                    "-f",
                    "lavfi",
                    "-i",
                    "color=c=black:s=320x180:d=2",
                    "-pix_fmt",
                    "yuv420p",
                    str(video),
                ],
                capture_output=True,
                text=True,
            )
            self.assertEqual(create_video.returncode, 0, create_video.stderr)
            normalized.write_text(
                json.dumps(
                    {
                        "events": [
                            {
                                "source_file": "sample.mp4",
                                "start_seconds": 0,
                                "end_seconds": 1,
                                "behavior": "证据不足测试",
                                "evidence_status": "证据不足",
                                "evidence_note": "黑色合成测试帧",
                                "evidence_frame_ids": ["sample-f000001"],
                            }
                        ]
                    },
                    ensure_ascii=False,
                ),
                encoding="utf-8",
            )
            result = subprocess.run(
                [
                    sys.executable,
                    str(PIPELINE_SCRIPT),
                    "--input",
                    str(video),
                    "--normalized-json",
                    str(normalized),
                    "--output",
                    str(output),
                ],
                capture_output=True,
                text=True,
            )
            self.assertEqual(result.returncode, 0, result.stderr)
            ws = load_workbook(output)["识别结果"]
            headers = header_map(ws)
            self.assertEqual(
                ws.cell(2, headers["证据帧标识"]).value,
                "sample.mp4@0.000s",
            )
            with zipfile.ZipFile(output) as archive:
                media = [
                    name for name in archive.namelist() if name.startswith("xl/media/")
                ]
            self.assertEqual(len(media), 1)


if __name__ == "__main__":
    unittest.main()
